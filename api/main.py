#!/usr/bin/env python
# -*- coding: utf-8 -*-
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse
from qdrant_client import QdrantClient, models
from qdrant_client.http.exceptions import UnexpectedResponse
import httpx
from fastembed import SparseTextEmbedding
import numpy as np
import json
import openpyxl
import io
import logging
import re
import asyncio
import os
import qdrant_client as qc
import fastembed
import fastapi
import importlib.metadata
from time import perf_counter
from typing import Optional
import uuid

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

logger.info("Starting API...")
try:
    logger.info(f"Qdrant Client version: {importlib.metadata.version('qdrant-client')}")
except Exception:
    logger.info("Qdrant Client version: unknown")
try:
    logger.info(f"FastEmbed version: {importlib.metadata.version('fastembed')}")
except Exception:
    logger.info("FastEmbed version: unknown")
try:
    logger.info(f"FastAPI version: {importlib.metadata.version('fastapi')}")
except Exception:
    logger.info("FastAPI version: unknown")

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/", response_class=HTMLResponse)
async def read_root():
    with open("static/index.html") as f:
        return f.read()

qdrant_client = QdrantClient(host=os.getenv("QDRANT_HOST", "qdrant"), port=int(os.getenv("QDRANT_PORT", "6333")))
upload_jobs = {}
stock_jobs = {}

def default_ollama_base_url():
    if os.path.exists("/.dockerenv"):
        return "http://ollama:11434"
    return "http://localhost:11434"

ollama_base_url = os.getenv("OLLAMA_BASE_URL", default_ollama_base_url())
ollama_api_url = f"{ollama_base_url}/api/embeddings"
ollama_batch_api_url = f"{ollama_base_url}/api/embed"
ollama_openai_embeddings_url = f"{ollama_base_url}/v1/embeddings"

# Initialize FastEmbed Sparse Model
# Using a standard SPLADE model which acts as a learned BM25 replacement.
sparse_embedding_model = SparseTextEmbedding(model_name="prithivida/Splade_PP_en_v1")

def safe_upsert(collection_name: str, points: list):
    try:
        qdrant_client.upsert(
            collection_name=collection_name,
            wait=True,
            points=points,
        )
        return
    except UnexpectedResponse as exc:
        message = str(exc)
        if "Payload error" in message and len(points) > 1:
            midpoint = len(points) // 2
            safe_upsert(collection_name, points[:midpoint])
            safe_upsert(collection_name, points[midpoint:])
            return
        raise


def create_collection(collection_name: str):
    try:
        qdrant_client.get_collection(collection_name=collection_name)
    except Exception:
        qdrant_client.recreate_collection(
            collection_name=collection_name,
            vectors_config={
                "text-dense": models.VectorParams(size=1024, distance=models.Distance.COSINE),
            },
            sparse_vectors_config={
                "text-sparse": models.SparseVectorParams(
                    index=models.SparseIndexParams(
                        on_disk=False,
                    )
                )
            },
        )

async def get_ollama_embeddings(texts, concurrency=4):
    async with httpx.AsyncClient(timeout=300.0) as client:
        # Try batch endpoint first (if supported by Ollama)
        try:
            batch_response = await client.post(
                ollama_batch_api_url,
                json={"model": "bge-m3", "input": texts},
            )
            if batch_response.status_code < 400:
                data = batch_response.json()
                if "embeddings" in data:
                    embeddings = data["embeddings"]
                    if len(embeddings) == len(texts):
                        return embeddings
                    logger.warning(
                        "Ollama /api/embed returned %s embeddings for %s inputs. Falling back.",
                        len(embeddings),
                        len(texts),
                    )
        except httpx.HTTPError:
            pass

        # Fallback: OpenAI-compatible embeddings endpoint (batch)
        try:
            openai_response = await client.post(
                ollama_openai_embeddings_url,
                json={"model": "bge-m3", "input": texts},
            )
            if openai_response.status_code < 400:
                data = openai_response.json()
                if "data" in data:
                    ordered = sorted(data["data"], key=lambda item: item.get("index", 0))
                    embeddings = [item["embedding"] for item in ordered]
                    if len(embeddings) == len(texts):
                        return embeddings
                    logger.warning(
                        "Ollama /v1/embeddings returned %s embeddings for %s inputs. Falling back.",
                        len(embeddings),
                        len(texts),
                    )
        except httpx.HTTPError:
            pass

        # Fallback: parallel single requests to /api/embeddings
        semaphore = asyncio.Semaphore(concurrency)

        async def fetch_one(text):
            async with semaphore:
                response = await client.post(
                    ollama_api_url,
                    json={"model": "bge-m3", "prompt": text},
                )
                if response.status_code == 404:
                    alt_response = await client.post(
                        ollama_openai_embeddings_url,
                        json={"model": "bge-m3", "input": text},
                    )
                    alt_response.raise_for_status()
                    data = alt_response.json()
                    return data["data"][0]["embedding"]
                response.raise_for_status()
                return response.json()["embedding"]

        try:
            return await asyncio.gather(*(fetch_one(text) for text in texts))
        except httpx.ConnectError as exc:
            logger.error(
                "Cannot connect to Ollama. Set OLLAMA_BASE_URL (e.g. http://ollama:11434 for Docker or http://localhost:11434 for local)."
            )
            raise exc

async def get_ollama_embedding(text: str):
    embeddings = await get_ollama_embeddings([text])
    return embeddings[0]

def get_sparse_embedding(text: str):
    # FastEmbed returns a generator of sparse embeddings
    embeddings = list(sparse_embedding_model.embed([text]))
    return embeddings[0]

async def process_xlsx_upload(
    *,
    contents: bytes,
    file_name: str,
    skip_rows: int,
    mappings: dict,
    collection_name: str,
    batch_size: int,
    points_batch_size: int,
    article_mode: str = "price",
    job_id: str = None,
):
    if points_batch_size <= 0:
        raise HTTPException(status_code=400, detail="points_batch_size must be > 0")

    workbook = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=skip_rows + 2, values_only=True))
    documents = []
    for row in rows:
        # Safely handle None values in row
        def get_val(idx):
            if idx < len(row) and row[idx] is not None:
                return str(row[idx])
            return ""

        text_to_embed = (
            f"Артикул - {get_val(mappings['Артикул'])}, "
            f"Наименование - {get_val(mappings['Наименование'])}, "
            f"Тариф с НДС, руб - {get_val(mappings['Тариф с НДС, руб'])}, "
            f"Имя файла - {file_name}"
        )
        documents.append(text_to_embed)

    total_start = perf_counter()
    indexed_rows = 0
    total_rows = len(documents)

    if job_id:
        upload_jobs[job_id].update(
            {
                "status": "running",
                "progress": 0,
                "indexed_rows": 0,
                "total_rows": total_rows,
                "rate": 0,
                "eta": None,
            }
        )

    # Batch dense embeddings via Ollama
    for start in range(0, len(documents), batch_size):
        batch_docs = documents[start : start + batch_size]
        batch_rows = rows[start : start + batch_size]

        embed_start = perf_counter()
        dense_task = asyncio.create_task(get_ollama_embeddings(batch_docs))
        sparse_vectors = list(sparse_embedding_model.embed(batch_docs))
        dense_vectors = await dense_task
        embed_duration = perf_counter() - embed_start

        batch_points = []
        for offset, dense_vector in enumerate(dense_vectors):
            i = start + offset
            sparse_vector = sparse_vectors[offset]

            qdrant_sparse_vector = models.SparseVector(
                indices=sparse_vector.indices.tolist(),
                values=sparse_vector.values.tolist(),
            )

            row_data = batch_rows[offset]
            payload = {}
            for header, col_idx in mappings.items():
                if col_idx < len(row_data):
                    payload[header] = row_data[col_idx]
            article_value = apply_article_mode(payload.get("Артикул"), article_mode)
            payload["Артикул"] = article_value
            payload["Остаток"] = 0
            payload["Имя файла"] = file_name
            point_id = build_point_id(collection_name, article_value)

            batch_points.append(
                models.PointStruct(
                    id=point_id,
                    vector={
                        "text-dense": dense_vector,
                        "text-sparse": qdrant_sparse_vector,
                    },
                    payload=payload,
                )
            )

        upsert_start = perf_counter()
        for chunk_start in range(0, len(batch_points), points_batch_size):
            chunk = batch_points[chunk_start : chunk_start + points_batch_size]
            safe_upsert(collection_name, chunk)
            indexed_rows += len(chunk)

            elapsed = perf_counter() - total_start
            rate = indexed_rows / elapsed if elapsed > 0 else 0
            remaining = total_rows - indexed_rows
            eta = remaining / rate if rate > 0 else 0
            percent = (indexed_rows / total_rows) * 100 if total_rows else 100

            logger.info(
                "Progress: %s/%s (%.1f%%), %.2f rows/sec, ETA %.1fs",
                indexed_rows,
                total_rows,
                percent,
                rate,
                eta,
            )

            if job_id:
                upload_jobs[job_id].update(
                    {
                        "status": "running",
                        "progress": round(percent, 2),
                        "indexed_rows": indexed_rows,
                        "total_rows": total_rows,
                        "rate": round(rate, 2),
                        "eta": round(eta, 1),
                    }
                )
        upsert_duration = perf_counter() - upsert_start

        logger.info(
            "Batch %s-%s: embeddings %.2fs, upsert %.2fs",
            start,
            start + len(batch_docs) - 1,
            embed_duration,
            upsert_duration,
        )

    total_duration = perf_counter() - total_start
    logger.info("Indexed %s rows in %.2fs", indexed_rows, total_duration)
    result = {
        "status": "success",
        "indexed_rows": indexed_rows,
        "duration_sec": round(total_duration, 3),
    }
    if job_id:
        upload_jobs[job_id].update(
            {
                "status": "completed",
                "progress": 100,
                "indexed_rows": indexed_rows,
                "total_rows": total_rows,
                "duration_sec": round(total_duration, 3),
            }
        )
    return result


@app.post("/upload_processed_xlsx")
async def upload_processed_xlsx(
    file: UploadFile = File(...),
    skip_rows: int = Form(...),
    mappings: str = Form(...),
    collection_name: str = Form(...),
    article_mode: str = Form("price"),
    batch_size: int = Form(16),
    points_batch_size: int = Form(200),
):
    logger.info(f"Received request to /upload_processed_xlsx for collection: {collection_name}")
    logger.info(
        f"skip_rows: {skip_rows}, mappings: {mappings}, batch_size: {batch_size}, points_batch_size: {points_batch_size}"
    )
    create_collection(collection_name)

    try:
        mappings = json.loads(mappings)
        if not mappings:
            raise HTTPException(status_code=400, detail="Mappings cannot be empty")
        logger.info(f"Parsed mappings: {mappings}")
        contents = await file.read()
        return await process_xlsx_upload(
            contents=contents,
            file_name=file.filename,
            skip_rows=skip_rows,
            mappings=mappings,
            collection_name=collection_name,
            batch_size=batch_size,
            points_batch_size=points_batch_size,
            article_mode=article_mode,
        )
    except Exception as e:
        logger.error(f"Error processing file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


async def run_upload_job(
    *,
    job_id: str,
    contents: bytes,
    file_name: str,
    skip_rows: int,
    mappings: dict,
    collection_name: str,
    batch_size: int,
    points_batch_size: int,
    article_mode: str,
):
    try:
        await process_xlsx_upload(
            contents=contents,
            file_name=file_name,
            skip_rows=skip_rows,
            mappings=mappings,
            collection_name=collection_name,
            batch_size=batch_size,
            points_batch_size=points_batch_size,
            article_mode=article_mode,
            job_id=job_id,
        )
    except Exception as exc:
        logger.error("Async upload job failed: %s", exc, exc_info=True)
        upload_jobs[job_id].update({"status": "failed", "error": str(exc)})


@app.post("/upload_processed_xlsx_async")
async def upload_processed_xlsx_async(
    file: UploadFile = File(...),
    skip_rows: int = Form(...),
    mappings: str = Form(...),
    collection_name: str = Form(...),
    article_mode: str = Form("price"),
    batch_size: int = Form(16),
    points_batch_size: int = Form(200),
):
    logger.info(f"Received request to /upload_processed_xlsx_async for collection: {collection_name}")
    logger.info(
        f"skip_rows: {skip_rows}, mappings: {mappings}, batch_size: {batch_size}, points_batch_size: {points_batch_size}"
    )
    create_collection(collection_name)

    try:
        mappings = json.loads(mappings)
        if not mappings:
            raise HTTPException(status_code=400, detail="Mappings cannot be empty")
        logger.info(f"Parsed mappings: {mappings}")
        contents = await file.read()
    except Exception as e:
        logger.error(f"Error processing file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

    job_id = str(uuid.uuid4())
    upload_jobs[job_id] = {"status": "queued", "progress": 0}
    asyncio.create_task(
        run_upload_job(
            job_id=job_id,
            contents=contents,
            file_name=file.filename,
            skip_rows=skip_rows,
            mappings=mappings,
            collection_name=collection_name,
            batch_size=batch_size,
            points_batch_size=points_batch_size,
            article_mode=article_mode,
        )
    )
    return {"status": "started", "job_id": job_id}


@app.get("/upload_status/{job_id}")
async def upload_status(job_id: str):
    if job_id not in upload_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    return upload_jobs[job_id]


def normalize_article(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.startswith('="') and text.endswith('"'):
        text = text[2:-1]
    if text.startswith("=") and text.endswith('"'):
        text = text[1:-1]
    if text.startswith('"') and text.endswith('"'):
        text = text[1:-1]
    return text.strip()


def apply_article_mode(article_value, article_mode: str) -> str:
    normalized = normalize_article(article_value)
    if not normalized:
        return normalized
    if article_mode == "chint":
        return f"{normalized}CHINT"
    if article_mode == "dkc":
        return f"DKC{normalized}"
    return normalized


def build_point_id(collection_name: str, article_value: str) -> str:
    if not article_value:
        return str(uuid.uuid4())
    return str(uuid.uuid5(uuid.NAMESPACE_URL, f"{collection_name}:{article_value}"))


def reset_stock_payload(collection_name: str):
    qdrant_client.set_payload(
        collection_name=collection_name,
        payload={"Остаток": 0},
        points=models.Filter(must=[]),
    )


async def process_stock_upload(
    *,
    contents: bytes,
    skip_rows: int,
    article_col: int,
    stock_col: int,
    collection_name: str,
    job_id: str = None,
    batch_size: int = 200,
):
    workbook = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=skip_rows + 2, values_only=True))
    total_rows = len(rows)
    processed = 0
    updated = 0
    skipped = 0
    total_start = perf_counter()

    if job_id:
        stock_jobs[job_id].update(
            {
                "status": "running",
                "progress": 0,
                "processed_rows": 0,
                "updated_rows": 0,
                "skipped_rows": 0,
                "total_rows": total_rows,
                "rate": 0,
                "eta": None,
            }
        )

    for start in range(0, total_rows, batch_size):
        batch_rows = rows[start : start + batch_size]
        batch_articles = []
        batch_payloads = {}

        for row in batch_rows:
            article = row[article_col] if article_col < len(row) else None
            stock_value = row[stock_col] if stock_col < len(row) else None
            if article is None:
                skipped += 1
                continue
            article_str = normalize_article(article)
            if not article_str:
                skipped += 1
                continue
            batch_articles.append(article_str)
            batch_payloads[article_str] = stock_value

        if not batch_articles:
            processed += len(batch_rows)
            continue

        matched_points = qdrant_client.scroll(
            collection_name=collection_name,
            scroll_filter=models.Filter(
                must=[
                    models.FieldCondition(
                        key="Артикул",
                        match=models.MatchAny(any=batch_articles),
                    )
                ]
            ),
            limit=len(batch_articles),
            with_payload=True,
            with_vectors=False,
        )[0]

        found_articles = set()
        for point in matched_points:
            article_value = point.payload.get("Артикул")
            if article_value is None:
                continue
            article_str = normalize_article(article_value)
            found_articles.add(article_str)
            if article_str not in batch_payloads:
                continue
            qdrant_client.set_payload(
                collection_name=collection_name,
                payload={"Остаток": batch_payloads.get(article_str)},
                points=[point.id],
            )
            updated += 1

        missing_articles = set(batch_articles) - found_articles
        skipped += len(missing_articles)

        processed += len(batch_rows)
        elapsed = perf_counter() - total_start
        rate = processed / elapsed if elapsed > 0 else 0
        remaining = total_rows - processed
        eta = remaining / rate if rate > 0 else 0
        percent = (processed / total_rows) * 100 if total_rows else 100

        logger.info(
            "Stock progress: %s/%s (%.1f%%), %.2f rows/sec, ETA %.1fs",
            processed,
            total_rows,
            percent,
            rate,
            eta,
        )

        if job_id:
            stock_jobs[job_id].update(
                {
                    "status": "running",
                    "progress": round(percent, 2),
                    "processed_rows": processed,
                    "updated_rows": updated,
                    "skipped_rows": skipped,
                    "total_rows": total_rows,
                    "rate": round(rate, 2),
                    "eta": round(eta, 1),
                }
            )

    duration = perf_counter() - total_start
    result = {
        "status": "success",
        "processed_rows": processed,
        "updated_rows": updated,
        "skipped_rows": skipped,
        "duration_sec": round(duration, 3),
    }

    if job_id:
        stock_jobs[job_id].update(
            {
                "status": "completed",
                "progress": 100,
                "processed_rows": processed,
                "updated_rows": updated,
                "skipped_rows": skipped,
                "duration_sec": round(duration, 3),
            }
        )

    return result


async def run_stock_job(
    *,
    job_id: str,
    contents: bytes,
    skip_rows: int,
    article_col: int,
    stock_col: int,
    collection_name: str,
    batch_size: int,
):
    try:
        stock_jobs[job_id].update({"status": "resetting", "progress": 0})
        reset_stock_payload(collection_name)
        await process_stock_upload(
            contents=contents,
            skip_rows=skip_rows,
            article_col=article_col,
            stock_col=stock_col,
            collection_name=collection_name,
            job_id=job_id,
            batch_size=batch_size,
        )
    except Exception as exc:
        logger.error("Stock upload job failed: %s", exc, exc_info=True)
        stock_jobs[job_id].update({"status": "failed", "error": str(exc)})


@app.post("/upload_stock_async")
async def upload_stock_async(
    file: UploadFile = File(...),
    skip_rows: int = Form(...),
    article_col: int = Form(...),
    stock_col: int = Form(...),
    collection_name: str = Form(...),
    batch_size: int = Form(200),
):
    logger.info(f"Received request to /upload_stock_async for collection: {collection_name}")
    logger.info(
        f"skip_rows: {skip_rows}, article_col: {article_col}, stock_col: {stock_col}, batch_size: {batch_size}"
    )

    try:
        contents = await file.read()
    except Exception as e:
        logger.error(f"Error reading stock file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

    job_id = str(uuid.uuid4())
    stock_jobs[job_id] = {"status": "queued", "progress": 0}
    asyncio.create_task(
        run_stock_job(
            job_id=job_id,
            contents=contents,
            skip_rows=skip_rows,
            article_col=article_col,
            stock_col=stock_col,
            collection_name=collection_name,
            batch_size=batch_size,
        )
    )
    return {"status": "started", "job_id": job_id}


@app.get("/stock_status/{job_id}")
async def stock_status(job_id: str):
    if job_id not in stock_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    return stock_jobs[job_id]

@app.get("/search")
async def search(
    query: str = Query(...),
    collection_name: str = Query(...),
    only_in_stock: bool = Query(False),
):
    # Dense vector for semantic search from Ollama
    dense_vector = await get_ollama_embedding(query)

    # Sparse vector from FastEmbed (No more re-indexing!)
    sparse_vector_gen = list(sparse_embedding_model.embed([query]))[0]
    sparse_vector = models.SparseVector(
        indices=sparse_vector_gen.indices.tolist(),
        values=sparse_vector_gen.values.tolist(),
    )

    def query_points(limit: int, query_filter: Optional[models.Filter]):
        return qdrant_client.query_points(
            collection_name=collection_name,
            prefetch=[
                models.Prefetch(
                    query=sparse_vector,
                    using="text-sparse",
                    limit=20,
                ),
                models.Prefetch(
                    query=dense_vector,
                    using="text-dense",
                    limit=20,
                ),
            ],
            query=models.FusionQuery(fusion=models.Fusion.RRF),
            limit=limit,
            with_payload=True,
            query_filter=query_filter,
        ).points

    # Hybrid search using Query API (Prefetch + Fusion RRF) - Requires qdrant-client >= 1.10.0
    try:
        in_stock_filter = models.Filter(
            must=[
                models.FieldCondition(
                    key="Остаток",
                    range=models.Range(gt=0),
                )
            ]
        )

        if only_in_stock:
            points = query_points(15, in_stock_filter)
            return {"results": points}

        in_stock_points = query_points(5, in_stock_filter)
        general_points = query_points(15, None)

        seen_ids = {point.id for point in in_stock_points}
        combined = list(in_stock_points)
        for point in general_points:
            if point.id in seen_ids:
                continue
            combined.append(point)
            seen_ids.add(point.id)
            if len(combined) >= 15:
                break

        return {"results": combined}
    except Exception as e:
        logger.error(f"Search failed: {e}", exc_info=True)
        # Return error details to the client for easier debugging
        raise HTTPException(status_code=500, detail=str(e))
